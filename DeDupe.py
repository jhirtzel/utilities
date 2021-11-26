#! python3

import openpyxl, sys 
from pprint import pprint

usage = """ Spreadsheet Deduplicator: takes a delimited data file as an input, and removes rows based on duplicates in the specified column.
	Usage: 'DeDupe <file> <column number> 
		- <file> should be a delimited text file or spreadsheet
		- <column number> should be an integer identifying which column holds duplicates, starting with column 1

		The program will search all rows in the specified column, and remove any rows after the first where a duplicate occurs
"""

if len(sys.argv) < 3:
	print(usage)
	sys.exit(1)
else:
	print(f'Removing duplicates based on column {sys.argv[2]}')

key_column = int(sys.argv[2])-1 #columns are 0-indexed, subtract one to get the nominal column
file = openpyxl.load_workbook(sys.argv[1],data_only=True)
sheet1 = file.active
deduped = {}


for row in sheet1.iter_rows(values_only=True):
	key = row[key_column]
	deduped[key] = row

file.create_sheet("dupes_removed")
sheet2 = file["dupes_removed"]
for row in deduped.values():
	sheet2.append(row)	
	
file.save(sys.argv[1])
		